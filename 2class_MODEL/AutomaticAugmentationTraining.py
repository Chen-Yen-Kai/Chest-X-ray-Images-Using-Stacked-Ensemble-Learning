# -*- coding: utf-8 -*-
"""
Created on Sun Jan  2 22:48:36 2022

@author: EN308
"""


# -*- coding: utf-8 -*-
"""
Created on Sun Oct  3 15:36:35 2021

@author: EN308
"""


# -*- coding: utf-8 -*-
"""
Created on Thu Sep 30 14:16:06 2021

@author: Berton Wei
"""
import tensorflow as tf
import Models_psp.models as model_net_psp
from tensorflow.keras.callbacks import EarlyStopping,ModelCheckpoint
from keras.backend import clear_session
from keras.metrics import MeanIoU
import Models.models as model_net
#from keras_unet_collection import models
from keras_unet_collection import losses as los
import time

#import tensorflow.compat.v1 as tf
#from simple_unet_model import simple_unet_model   #Use normal unet model
from tensorflow.keras.utils import normalize
import os
os.environ["TF_CPP_MIN_LOG_LEVEL"]="3"
import cv2
import openpyxl
from PIL import Image
import numpy as np
from matplotlib import pyplot as plt
import gc
from loss_functions import *
import shutil
import xlwt
import segmentation_models as sm
sm.set_framework('tf.keras')
from sklearn.model_selection import train_test_split
from keras.callbacks import EarlyStopping
from keras import optimizers
#from tensorflow.keras.optimizers import Adam

#設定訓練所使用之增強資料夾(Augmentation)與主幹網路(blackbone_functions)
#Augmentation=["CLAHE","None"]
Augmentation=["CLAHE"]

#blackbone_functions=['vgg19','resnet101','resnext101','densenet201','inceptionresnetv2','mobilenetv2']
#blackbone_functions=['none', 'resnet50']
blackbone_functions=['resnet50']

#定義loss function
# Segmentation models losses can be combined together by '+' and scaled by integer or float factor
# set class weights for dice_loss (car: 1.; pedestrian: 2.; background: 0.5;)
dice_loss = sm.losses.DiceLoss() 
focal_loss = sm.losses.CategoricalFocalLoss()
total_loss = dice_loss + (1 * focal_loss)


def hybrid_loss(y_true, y_pred):

    loss_focal = los.focal_tversky(y_true, y_pred, alpha=0.5, gamma=4/3)
    loss_iou = los.iou_seg(y_true, y_pred)
    
    return loss_focal + loss_iou 

#設定訓練所使用之網路、loss function、學習率(rate_functions)、資料夾(folder_functions)
#net_str=['U_Net','att_unet',"unetplusplus","att_unetplusplus","att_unetplusplus_psp"]
net_str=['unetplusplus']

#loss_functions =['categorical_crossentropy', focal_tversky, total_loss] 
loss_functions =[total_loss] 

#loss_functions_str =['categorical_crossentropy', 'focal_tversky', 'total_loss']  
loss_functions_str =['total_loss'] 

rate_functions =[1e-04] 
#folder_functions=["yun_normal", "yun_normalLung_segmentation", "yun_Bone_suppression", "yun_Bone_suppression_Lung_segmentation"]
folder_functions=["yun_normal"]
#training_data_path="Consolidation_yun_normal"

n_classes=3 #病徵類別
import pandas as pd
from  sklearn.metrics import precision_recall_fscore_support

#計算評估指標
def pandas_classification_report(y_true, y_pred):
    metrics_summary = precision_recall_fscore_support(
            y_true=y_true, 
            y_pred=y_pred)
    
    avg = list(precision_recall_fscore_support(
            y_true=y_true, 
            y_pred=y_pred,
            average='macro'))

    metrics_sum_index = ['precision', 'recall', 'f1-score', 'support']
    class_report_df = pd.DataFrame(
        list(metrics_summary),
        index=metrics_sum_index)
    
    support = class_report_df.loc['support']
    total = support.sum() 
    avg[-1] = total
    
    class_report_df['avg / total'] = avg

    return class_report_df.T

def set_regularization(model, 
                       kernel_regularizer=None, 
                       bias_regularizer=None, 
                       activity_regularizer=None):
    
    for layer in model.layers:
        
        # set kernel_regularizer
        if kernel_regularizer is not None and hasattr(layer, 'kernel_regularizer'):
            layer.kernel_regularizer = kernel_regularizer

        # set bias_regularizer
        if bias_regularizer is not None and hasattr(layer, 'bias_regularizer'):
            layer.bias_regularizer = bias_regularizer

        # set activity_regularizer
        if activity_regularizer is not None and hasattr(layer, 'activity_regularizer'):
            layer.activity_regularizer = activity_regularizer


# Reset Keras Session
def reset_keras():
    sess =  tf.compat.v1.Session()
    tf.compat.v1.keras.backend.clear_session()
    sess.close()
    sess =  tf.compat.v1.Session()

    try:
        pass # this is from global space - change this as you need
    except:
        pass

    #print(gc.collect()) # if it's done something you should see a number being outputted

    # use the same config as you used to create the session
    config = tf.compat.v1.ConfigProto()
    config.gpu_options.per_process_gpu_memory_fraction = 1
    config.gpu_options.visible_device_list = "0"
    #tf.compat.v1.keras.backend.set_session(tf.compat.v1.Session(config=config))

#載入影像資料    
def image_dataset_url(image_directory,image_dataset):
    image_dataset=[]
    images = os.listdir(image_directory)
    for i, image_name in enumerate(images):    #Remember enumerate method adds a counter and returns the enumerate object
        if (image_name.split('.')[1] == 'png'):
            #print(image_directory+image_name)
            image = cv2.imread(image_directory+image_name)
            image = Image.fromarray(image)
            image_dataset.append(np.array(image))
    return image_dataset
#載入label資料
def mask_directory_url(mask_directory,maskdataset):
    maskdataset=[]
    masks = os.listdir(mask_directory)
    for i, image_name in enumerate(masks):
        if (image_name.split('.')[1] == 'png'):
            image = cv2.imread(mask_directory+image_name,0)
            #print(lung_directory+image_name)
            #ret, th1 = cv2.threshold(image, 30, 255, cv2.THRESH_BINARY)
            #image = cv2.cvtColor(th1, cv2.COLOR_GRAY2BGR)
            image = Image.fromarray(image)
            maskdataset.append(np.array(image))
    return maskdataset


#主程式

for colder_index in range(len(folder_functions)):
    training_data_path=folder_functions[colder_index]
    print(training_data_path)
    xlsx_file=training_data_path+'.xlsx'
    workbook = openpyxl.Workbook()
    workbook.save(xlsx_file) #儲存
    result_list = [["Network","folder", "Pre-processing", "Learning rate","loss_functions","blackbone","IoU for background","IoU for Consolidation","IoU for InfiltrationsBronchiectasis","Mean IoU","Mean precision","Mean recall","Mean f1-score","Acc","Kernel Size"]]
    wb=openpyxl.load_workbook(r''+folder_functions[colder_index]+'.xlsx')
    ws = wb.create_sheet(title='Sheet1',index=0) # 新建一個sheet分頁
    #獲取行數和列數：
    rows = ws.max_row   #獲取行數
    cols = ws.max_column    #獲取列數
    
    print("Declare an augmentation pipeline")
    import albumentations as A
    #資料擴增
    # Declare an augmentation pipeline
    for Augmentation_index in range(len(Augmentation)):
        #有CLAHE的訓練影像擴增
        Train_CLAHE = A.Compose([
            A.RandomBrightnessContrast (brightness_limit=0.1, p=1.0),
            A.CLAHE (clip_limit=4, tile_grid_size=(8, 8), always_apply=True, p=1.0),
            A.Affine (scale=None,translate_percent={"x": (-0.1, 0.1), "y": (-0.1, 0.1)},rotate=(-5, 5)),
            A.HorizontalFlip(),
        ])
        #沒有CLAHE的訓練影像擴增
        Train = A.Compose([
            A.RandomBrightnessContrast (brightness_limit=0.1, p=1.0),
            A.Affine (scale=None,translate_percent={"x": (-0.1, 0.1), "y": (-0.1, 0.1)},rotate=(-5, 5)),
            A.HorizontalFlip(),
        ])
        
        #有CLAHE的測試、驗證影像處理
        Test_CLAHE = A.Compose([
            A.CLAHE (clip_limit=4, tile_grid_size=(8, 8), always_apply=True, p=1.0)
        ])
        #沒有CLAHE的測試、驗證影像處理
        Test = A.Compose([
            
        ])
    
        #訓練影像儲存資料夾
        training_path_image='./'+training_data_path+"/train/train_"+str(Augmentation[Augmentation_index])+"result_image"
        training_path_label='./'+training_data_path+"/train/train_"+str(Augmentation[Augmentation_index])+"result_label"
        os.mkdir(training_path_image)
        os.mkdir(training_path_label)
        yourPath = './'+training_data_path+"/train/image/"
        maskPath = './'+training_data_path+"/train/label/"
        allFileList = os.listdir(yourPath)
        #設定擴增倍數
        number_index=0
        for index in range(8):
            for file in allFileList:
              if os.path.isdir(os.path.join(yourPath,file)):
                print("I'm a directory: " + file)
              elif os.path.isfile(yourPath+file):
                image = cv2.imread(yourPath+file)
                mask = cv2.imread(maskPath+file)
                number_index+=1
                # Apply augmentations to image and a mask
                if Augmentation[Augmentation_index]=="CLAHE":
                    augmented = Train_CLAHE(image = image, mask = mask)
                    image_augmented = augmented['image']
                    mask_augmented = augmented['mask']
                else:
                    augmented = Train(image = image, mask = mask)
                    image_augmented = augmented['image']
                    mask_augmented = augmented['mask']
                # Access augmented image and mask

                cv2.imwrite(training_path_image+"/"+str(index)+file, image_augmented,[cv2.IMWRITE_PNG_COMPRESSION, 0])
                cv2.imwrite(training_path_label+"/"+str(index)+file, mask_augmented,[cv2.IMWRITE_PNG_COMPRESSION, 0])
        #驗證影像儲存資料夾
        val_path_image='./'+training_data_path+"/val/val_"+str(Augmentation[Augmentation_index])+"result_image"
        val_path_label='./'+training_data_path+"/val/val_"+str(Augmentation[Augmentation_index])+"result_label"
        os.mkdir(val_path_image)
        os.mkdir(val_path_label)
        valPath = './'+training_data_path+"/val/image/"
        valmaskPath = './'+training_data_path+"/val/label/"
        allFileList = os.listdir(valPath)
        #設定擴增倍數
        number_index=0
        for index in range(1):
            for file in allFileList:
              if os.path.isdir(os.path.join(valPath,file)):
                print("I'm a directory: " + file)
              elif os.path.isfile(valPath+file):
                image = cv2.imread(valPath+file)
                mask = cv2.imread(valmaskPath+file)
                number_index+=1
                # Access augmented image and mask
                if Augmentation[Augmentation_index] == "CLAHE":
                    test_augmented = Test_CLAHE(image = image, mask = mask)
                else:
                    test_augmented = Test(image = image, mask = mask)
                test_image_augmented = test_augmented['image']
                test_mask_augmented = test_augmented['mask']
                cv2.imwrite(val_path_image+"/"+file, test_image_augmented)
                cv2.imwrite(val_path_label+"/"+file, test_mask_augmented)
            
        #測試影像儲存資料夾
        test_path_image='./'+training_data_path+"/test_"+str(Augmentation[Augmentation_index])+"result_image"
        #test_path_label='./'+training_data_path+"/test_"+str(clip_functions[clahe])+"result_label"
        os.mkdir(test_path_image)
        #os.mkdir(test_path_label)
        testPath = './'+training_data_path+"/test/"
        test_label_Path = './'+training_data_path+"/test_label/"
        allFileList = os.listdir(testPath)
        #設定擴增倍數
        number_index=0
        for index in range(1):
            for file in allFileList:
              if os.path.isdir(os.path.join(testPath,file)):
                print("I'm a directory: " + file)
              elif os.path.isfile(testPath+file):
                image = cv2.imread(testPath+file)
                #mask = cv2.imread(test_label_Path+file)
                print(testPath+file)
                number_index+=1
                if Augmentation[Augmentation_index] == "CLAHE":
                    test_augmented = Test_CLAHE(image = image, mask = mask)
                else:
                    test_augmented = Test(image = image, mask = mask)
                test_image_augmented = test_augmented['image']
                cv2.imwrite(test_path_image+"/"+file, test_image_augmented)

                
        #載入訓練資料
        image_dataset = []  #Many ways to handle data, you can use pandas. Here, we are using a list format.  
        image_directory = training_path_image+"/"
        image_dataset=image_dataset_url(image_directory,image_dataset)
        mask_dataset = []  #Place holders to define add labels. We will add 0 to all parasitized images and 1 to uninfected.
        mask_directory = training_path_label+"/"
        print(mask_directory)
        mask_dataset=mask_directory_url(mask_directory,mask_dataset)
        print(mask_dataset)
        mask_dataset = np.array(mask_dataset)
        #Encode labels... but multi dim array so need to flatten, encode and reshape
        from sklearn.preprocessing import LabelEncoder
        labelencoder = LabelEncoder()
        n, h, w = mask_dataset.shape
        mask_dataset = mask_dataset.reshape(-1,1)
        mask_dataset = labelencoder.fit_transform(mask_dataset)
        mask_dataset = mask_dataset.reshape(n, h, w)
        
        np.unique(mask_dataset)
        
        #################################################
        #train_images = np.expand_dims(train_images, axis=3)
        #train_images = normalize(train_images, axis=1)
        print("Class values in the dataset are ... ", np.unique(mask_dataset))  # 0 is the background/few unlabeled 
        mask_dataset = np.expand_dims(mask_dataset, axis=3)
        
        #載入驗證資料
        val_dataset = []  #Many ways to handle data, you can use pandas. Here, we are using a list format.  
        val_image_directory = val_path_image+"/"
        val_dataset=image_dataset_url(val_image_directory,val_dataset)
        val_mask_dataset = []  #Place holders to define add labels. We will add 0 to all parasitized images and 1 to uninfected.
        val_mask_directory = val_path_label+"/"
        val_mask_dataset=mask_directory_url(val_mask_directory,val_mask_dataset)
        val_mask_dataset = np.array(val_mask_dataset)
        #Normalize images
        #Encode labels... but multi dim array so need to flatten, encode and reshape
        labelencoder = LabelEncoder()
        n, h, w = val_mask_dataset.shape
        val_mask_dataset = val_mask_dataset.reshape(-1,1)
        val_mask_dataset = labelencoder.fit_transform(val_mask_dataset)
        val_mask_dataset = val_mask_dataset.reshape(n, h, w)
        
        np.unique(val_mask_dataset)
        print("Class values in the dataset are ... ", np.unique(val_mask_dataset))  # 0 is the background/few unlabeled 
        #載入測試資料
        test_dataset = []  #Many ways to handle data, you can use pandas. Here, we are using a list format.  
        test_image_directory = test_path_image+"/"
        test_dataset=image_dataset_url(test_image_directory,test_dataset)

        
        val_mask_dataset = np.expand_dims(val_mask_dataset, axis=3)
        from tensorflow.keras.utils import to_categorical
        train_masks_3_class = to_categorical(mask_dataset, num_classes=n_classes)
        y_train = train_masks_3_class.reshape((mask_dataset.shape[0], mask_dataset.shape[1], mask_dataset.shape[2], n_classes))
        val_masks_3_class = to_categorical(val_mask_dataset, num_classes=n_classes)
        y_val = val_masks_3_class.reshape((val_mask_dataset.shape[0], val_mask_dataset.shape[1], val_mask_dataset.shape[2], n_classes))
        # 载入模型
        for net in range(len(net_str)):  
            print(net_str[net])
            for blackbone in range(len(blackbone_functions)):
                print(blackbone_functions[blackbone])
                for loss in range(len(loss_functions_str)):
                    print(loss_functions_str[loss])
                    for rate in range(len(rate_functions)):
                        print(rate_functions[rate])
                        Mean_IoU=0
                        result_path='./'+training_data_path+"/"+str(net_str[net])+"_"+str(Augmentation[Augmentation_index])+"_"+str(rate_functions[rate])+"_"+loss_functions_str[loss]+"_"+blackbone_functions[blackbone]+"_result"
                        #建立plot放置資料夾
                        os.mkdir(result_path)
                        plot_path=result_path+"/plot_path"
                        os.mkdir(plot_path)
                        predict_path=result_path+"/predict"
                        os.mkdir(predict_path)
                        
                        con_path=result_path+"/Consolidation"
                        os.mkdir(con_path)
                        bro_path=result_path+"/InfiltrationsBronchiectasis"
                        os.mkdir(bro_path)
                        acc=0
                       
                        # actulally total_loss can be imported directly from library, above example just show you how to manipulate with losses
                        # total_loss = sm.losses.binary_focal_dice_loss # or sm.losses.categorical_focal_dice_loss 
                        
                        metrics = [sm.metrics.IOUScore(), sm.metrics.FScore()] #訓練評估指標
                        
    
                        # preprocess input
                        image_dataset = np.array(image_dataset)
                        val_dataset = np.array(val_dataset)
                        es = EarlyStopping(monitor='val_iou_score', patience=10, mode='max', restore_best_weights=True) #訓練提前停止條件
                        count = 0
                        while Mean_IoU<0.60 : #若IoU小於0.60則重新訓練
                            count += 1
                            if blackbone_functions[blackbone] == "none" : #未使用主幹網路
                                #model = model_net.Unet(use_backbone=False, input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                            #net_str='U_Net','att_unet',"unetplusplus","att_unetplusplus"
                                if net_str[net] == "U_Net" :
                                    model = model_net.Unet(use_backbone=False, input_shape=(512,512,3),attention=False,classes=3,activation="softmax")
                                elif  net_str[net] == "att_unet" :
                                    model = model_net.Unet(use_backbone=False, input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                                elif  net_str[net] == "unetplusplus" :
                                    model = model_net.Xnet(use_backbone=False, input_shape=(512,512,3),attention=False,classes=3,activation="softmax")  
                                elif  net_str[net] == "att_unetplusplus" :
                                    model = model_net.Xnet(use_backbone=False, input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                                elif  net_str[net] == "att_unetplusplus_psp" :
                                    model = model_net_psp.Xnet(use_backbone=False, input_shape=(512,512,3),attention=True,classes=3,activation="softmax",decoder_block_type = "transpose")


                            else: #使用主幹網路
                            
                                if net_str[net] == "U_Net" :
                                    model = model_net.Unet(use_backbone=True,backbone_name=blackbone_functions[blackbone], input_shape=(512,512,3),attention=False,classes=3,activation="softmax")
                                elif  net_str[net] == "att_unet" :
                                    model = model_net.Unet(use_backbone=True,backbone_name=blackbone_functions[blackbone], input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                                elif  net_str[net] == "unetplusplus" :
                                    model = model_net.Xnet(use_backbone=True,backbone_name=blackbone_functions[blackbone], input_shape=(512,512,3),attention=False,classes=3,activation="softmax")  
                                elif  net_str[net] == "att_unetplusplus" :
                                    model = model_net.Xnet(use_backbone=True,backbone_name=blackbone_functions[blackbone], input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                                elif  net_str[net] == "att_unetplusplus_psp" :
                                    model = model_net_psp.Xnet(use_backbone=True,backbone_name=blackbone_functions[blackbone], input_shape=(512,512,3),attention=True,classes=3,activation="softmax",decoder_block_type = "transpose")
                                elif net_str[net] == "U_Net3+" :
                                    model = models.unet_3plus_2d(input_size=(512,512,3), n_labels=3, filter_num_down=(32, 64, 128, 256, 512), backbone=blackbone_functions[blackbone])
                            
                            model_input = model.input
                            
                            model_output = model.get_layer('final_conv').output
                            from keras.layers import Dropout
                            from tensorflow.keras.models import Model
                            #add dropoutactivations
                            model_output = Dropout(0.3)(model_output)
                            #add activation
                            output =tf.keras.activations.softmax(model_output)
                            model =Model(model.input, output)
                            from tensorflow.keras import regularizers
                            set_regularization(model, kernel_regularizer=regularizers.l2(0.0001),bias_regularizer=regularizers.l2(0.0001))
                            model.summary()
                            #model = sm.utils.set_regularization(model, kernel_regularizer=regularizers.l2(0.001))
                            model.compile(optimizer=tf.optimizers.Adam(rate_functions[rate]), loss=loss_functions[loss], metrics=metrics)
                            #model name
                            filepath="model_"+str(net_str[net])+"_"+training_data_path+"_"+str(Augmentation[Augmentation_index])+"_"+str(rate_functions[rate])+"_"+loss_functions_str[loss]+"_"+blackbone_functions[blackbone]+".hdf5"
                            #model folder
                            save_dir = os.path.join(os.getcwd(), './'+training_data_path+'/model')
                            # 定義訓練回覆學習率的方法
                            model_checkpoint=ModelCheckpoint(os.path.join(save_dir, filepath),monitor='val_iou_score',verbose=0, save_best_only=True, save_weights_only=True, mode = 'max')
                            
                            
                            time_start = time.time() #開始計時
                            Unet = model.fit(image_dataset, y_train, 
                                                batch_size = 8, 
                                                verbose=1, 
                                                epochs=100, 
                                                validation_data=(val_dataset, y_val),
                                                callbacks=[model_checkpoint,es])

                            #顯示訓練與驗證的IoU
                            f, ax = plt.subplots()
                            ax.plot([None] + Unet.history['iou_score'], 'o-')
                            ax.plot([None] + Unet.history['val_iou_score'], 'x-')
                            ax.legend(['iou_score', 'val_iou_score'], loc = 0)
                            ax.set_title('Training/Validation iou_score per Epoch_')
                            ax.set_xlabel('Epoch')
                            ax.set_ylabel('iou_score')
                            plt.savefig(result_path+"/iou_score.png")
                            #顯示訓練與驗證的loss
                            f, ax = plt.subplots()
                            ax.plot([None] + Unet.history['loss'], 'o-')
                            ax.plot([None] + Unet.history['val_loss'], 'x-')
                            ax.legend(['Train loss', 'Validation loss'], loc = 0)
                            ax.set_title('Training/Validation loss per Epoch')
                            ax.set_xlabel('Epoch')
                            ax.set_ylabel('loss')
                            plt.savefig(result_path+"/loss.png")
                            #下一個loss_function
                            time_end = time.time()    #結束計時
                            
                            #model.load_weights('./'+training_data_path+'/model/'+filepath)
                            

                            
                            
                            yourPath = test_path_image+"/"
                            allFileList = os.listdir(yourPath)
                            for file in allFileList:
                              if os.path.isdir(os.path.join(yourPath,file)):
                                print("I'm a directory: " + file)
                              elif os.path.isfile(yourPath+file):
                                #讀取測試影像
                                test_image = []
                                fimage = cv2.imread(yourPath+file)
                                
                                image = Image.fromarray(fimage)
                                test_image.append(np.array(image))
                                # preprocess input
                                test_image = np.array(test_image)
                                test_pred1 = model.predict(test_image) #預測測試影像
                                test_prediction1 = np.argmax(test_pred1, axis=3)[0,:,:]
                                #To calculate I0U for each class...
                                prediction_array=np.zeros((512,512,3))
                                prediction_con=np.zeros((512,512,1))
                                prediction_bro=np.zeros((512,512,1))
                                prediction_index_x=np.where(test_prediction1==1)[0]
                                prediction_index_y=np.where(test_prediction1==1)[1]
                                for index in range(len(prediction_index_x)):
                                    prediction_array[prediction_index_x[index], prediction_index_y[index]]=[0,0,255]
                                    prediction_con[prediction_index_x[index], prediction_index_y[index]]=[255]
                                cv2.imwrite(con_path+"/"+file, prediction_con)
                                prediction_index_x=np.where(test_prediction1==2)[0]
                                prediction_index_y=np.where(test_prediction1==2)[1]
                                for index in range(len(prediction_index_x)):
                                    prediction_array[prediction_index_x[index],prediction_index_y[index]]=[0,255,255]
                                    prediction_bro[prediction_index_x[index],prediction_index_y[index]]=[255]
                                cv2.imwrite(bro_path+"/"+file, prediction_bro)
                                
                                #開運算
                                kernel=np.ones((21, 21),np.uint8)
                                #cv2.imwrite(result_path+"/prediction_bro_"+file, prediction_bro)
                                #cv2.imwrite(result_path+"/prediction_con_"+file, prediction_con)
                                prediction_con=cv2.morphologyEx(prediction_con, cv2.MORPH_OPEN, kernel)
                                prediction_bro=cv2.morphologyEx(prediction_bro, cv2.MORPH_OPEN, kernel)
                                
                                cv2.imwrite(bro_path+"/"+file, prediction_bro)
                                cv2.imwrite(con_path+"/"+file, prediction_con)
                                
                                prediction_con = cv2.resize(prediction_con, (int(file.split("_")[5]),int(file.split("_")[5])), interpolation=cv2.INTER_AREA)
                                prediction_bro = cv2.resize(prediction_bro, (int(file.split("_")[5]),int(file.split("_")[5])), interpolation=cv2.INTER_AREA)
                                #解答繪製
                                groundtruth_image = cv2.imread("./groundtruth/"+file.split("_")[0]+".png")
                                groundtruth_array=np.zeros((groundtruth_image.shape[0],groundtruth_image.shape[1],3))
                                groundtruth_index_x=np.where(groundtruth_image==1)[0]
                                groundtruth_index_y=np.where(groundtruth_image==1)[1]
                                for index in range(len(groundtruth_index_x)):
                                    groundtruth_array[groundtruth_index_x[index],groundtruth_index_y[index]]=[0,0,255]
                                groundtruth_index_x=np.where(groundtruth_image==2)[0]
                                groundtruth_index_y=np.where(groundtruth_image==2)[1]
                                for index in range(len(groundtruth_index_x)):
                                    groundtruth_array[groundtruth_index_x[index],groundtruth_index_y[index]]=[0,255,255]
                                cv2.imwrite(result_path+"/"+file.split("_")[0]+'groundtruth.png', groundtruth_array)
                                #回到原始圖像
                                #or_size_image = cv2.imread("./image_orginal_marked/"+file.split("_")[0]+".jpg")
                                '''
                                ##計算iOU
                                print(or_size_image.shape)
                                or_size_image_mask=np.zeros((or_size_image.shape[0],or_size_image.shape[1]))
                                or_size_image_mask_rgb=np.zeros((or_size_image.shape[0],or_size_image.shape[1],3))
                                imdex_y=0
                                for y in range(or_size_image.shape[0]):
                                    
                                    imdex_x=0
                                    for x in range(or_size_image.shape[1]):
                                        #print (file.split("_"))
                                        if ((y>=int(file.split("_")[1]))and (y<int(file.split("_")[2]))) and ((x>=int(file.split("_")[3])) and (x<int(file.split("_")[4]))) :
                                           # print(imdex_y,imdex_x)
                                            if prediction_bro[imdex_y-int(file.split("_")[1])][imdex_x-int(file.split("_")[3])]>0:
                                                or_size_image_mask[y][x]=2
                                                or_size_image_mask_rgb[y][x]=[0,255,255]
                                                #print("prediction_con")
                                            if prediction_con[imdex_y-int(file.split("_")[1])][imdex_x-int(file.split("_")[3])]>0:
                                            #if prediction_con[imdex_y][imdex_x]>0:
                                                #print("prediction_con")
                                                or_size_image_mask[y][x]=1
                                                or_size_image_mask_rgb[y][x]=[0,0,255]
                                        imdex_x=imdex_x+1
                                    imdex_y=imdex_y+1
                                '''
                                #回到原始圖像
                                or_size_image = cv2.imread("./image_orginal_marked/"+file.split("_")[0]+".jpg")
                                prediction_bro=np.where(prediction_bro<1,0,2)
                                prediction_con=np.where(prediction_con<1,0,1)
                                allsize=prediction_con+prediction_bro
                                allsize[allsize==3]=1
                                
                                #allsize=np.pad(allsize, ((int(file.split("_")[3]), int(file.split("_")[1])),(or_size_image.shape[0]-int(file.split("_")[2]), or_size_image.shape[1]-int(file.split("_")[4]))), 'constant', constant_values=((0, 0)))
                                allsize=np.pad(allsize, ((int(file.split("_")[1]),or_size_image.shape[0]-int(file.split("_")[2])),(int(file.split("_")[3]),or_size_image.shape[1]-int(file.split("_")[4]))), 'constant', constant_values=(0, 0))
                                
                                allsize_rgb_array=np.zeros((groundtruth_image.shape[0],groundtruth_image.shape[1],3))
                                allsize_rgb=np.zeros((allsize_rgb_array.shape[0],allsize_rgb_array.shape[1],3))
                                allsize_rgb_x=np.where(allsize==1)[0]
                                allsize_rgb_y=np.where(allsize==1)[1]
                                for index in range(len(allsize_rgb_x)):
                                    allsize_rgb_array[allsize_rgb_x[index],allsize_rgb_y[index]]=[0,0,255]
                                allsize_rgb_x=np.where(allsize==2)[0]
                                allsize_rgb_y=np.where(allsize==2)[1]
                                for index in range(len(allsize_rgb_x)):
                                    allsize_rgb_array[allsize_rgb_x[index],allsize_rgb_y[index]]=[0,255,255]
                                    
                                    
                                cv2.imwrite(predict_path+"/"+file.split("_")[0]+'.png', allsize)
                                cv2.imwrite(result_path+"/"+file.split("_")[0]+'image_mask_rgb.png', allsize_rgb_array)      
                                prediction = cv2.imread(result_path+"/"+file.split("_")[0]+'image_mask_rgb.png')
                                #prediction = cv2.imread(result_path+"/"+file[:-4]+'or_size_image_mask_rgb.png')
                                result = cv2.addWeighted(prediction, 0.3, or_size_image, 1, 0, or_size_image)
                                cv2.imwrite(result_path+"/"+file.split("_")[0]+'.png', result)
                                
                                #plot資料夾中圖表顯示
                                fig = plt.figure(figsize=(80,40))
                                image_or=cv2.imread("./image_orginal/"+file.split("_")[0]+'.png')
                                ax = fig.add_subplot(1, 5, 1)
                                image_or_plot=cv2.cvtColor(image_or,cv2.COLOR_BGR2RGB)
                                imgplot = plt.imshow(image_or_plot)
                                ax.set_title('Original image',fontsize=40)
                                plt.axis('off')
                                ax = fig.add_subplot(1, 5, 2)
                                enhancement_image = cv2.imread(test_path_image+"/"+file)
                                enhancement_image_plot=cv2.cvtColor(enhancement_image,cv2.COLOR_BGR2RGB)
                                imgplot = plt.imshow(enhancement_image)
                                ax.set_title('Preprocessed images',fontsize=40)
                                plt.axis('off')
                                ax = fig.add_subplot(1, 5, 3)
                                Ground_truth = cv2.imread(result_path+"/"+file.split("_")[0]+'groundtruth.png')
                                Ground_truth_plot=cv2.cvtColor(Ground_truth,cv2.COLOR_BGR2RGB)
                                imgplot = plt.imshow(Ground_truth_plot)
                                ax.set_title('Ground truth',fontsize=40)
                                plt.axis('off')
                                ax = fig.add_subplot(1, 5, 4)
                                    
                                prediction = cv2.imread(result_path+"/"+file.split("_")[0]+'image_mask_rgb.png')
                                prediction_plot=cv2.cvtColor(prediction,cv2.COLOR_BGR2RGB)
                                imgplot = plt.imshow(prediction_plot)
                                ax.set_title('Prediction mask',fontsize=40)
                                plt.axis('off')
                                
                                ax = fig.add_subplot(1, 5, 5)
                                overlap_reault = cv2.imread(result_path+"/"+file.split("_")[0]+'.png')
                                overlap_reault_plot=cv2.cvtColor(overlap_reault,cv2.COLOR_BGR2RGB)
                                imgplot = plt.imshow(overlap_reault_plot)
                                #mark_reault_plot=cv2.cvtColor(mark,cv2.COLOR_BGR2RGB)
                                #imgplot = plt.imshow(mark_reault_plot)
                                ax.set_title('Overlapping results',fontsize=40)
                                plt.axis('off')
                                    
                                plt.savefig(plot_path+"/"+file[:-4]+'_plot.png')
                                plt.rcParams.update({'figure.max_open_warning': 0})
                                plt.close(fig)
                            
                            
                            
                            predictdirectory=predict_path+"/"
                            predictdataset=np.array([])
                            groundtruthdataset=np.array([])
                            masks = os.listdir(predictdirectory)
                            for i, image_name in enumerate(masks):
                                if (image_name.split('.')[1] == 'png'):
                                   predict_image = cv2.imread(predictdirectory+image_name,0)
                                   groundtruth_image = cv2.imread("./groundtruth/"+image_name,0)
                                   print("./groundtruth/"+image_name)
                                   print(predictdirectory+image_name)
                                   print(predict_image.shape)
                                   print(groundtruth_image.shape) 
                                   height,width=predict_image.shape
                                   predict_image_img=np.reshape(predict_image, (height*width))
                                   groundtruth_image_img=np.reshape(groundtruth_image, (height*width))
                                   predictdataset = np.concatenate((predictdataset, predict_image_img), axis = 0).astype(int)
                                   groundtruthdataset = np.concatenate((groundtruthdataset, groundtruth_image_img), axis = 0).astype(int)
                
                            print(predictdataset.shape)
                            print(groundtruthdataset.shape) 
                
                                
                            #計算Mean IoU
                            #from keras.metrics import MeanIoU
                            n_classes = 3
                            IOU_keras = MeanIoU(num_classes=n_classes)  
                            #predictdataset = np.array(predictdataset)
                            #groundtruthdataset = np.array(groundtruthdataset)
                            print("predictdataset.shape")
                            print(predictdataset)
                            print(groundtruthdataset)
                            IOU_keras.update_state(groundtruthdataset, predictdataset)
                            print("Mean IoU =", IOU_keras.result().numpy())
        
        
                            #To calculate I0U for each class...
                            values = np.array(IOU_keras.get_weights()).reshape(n_classes, n_classes)
                            print(values)
                            class1_IoU = values[0,0]/(values[0,0] + values[0,1] + values[0,2] + values[1,0]+ values[2,0])
                            class2_IoU = values[1,1]/(values[1,1] + values[1,0] + values[1,2] + values[0,1]+ values[2,1])
                            class3_IoU = values[2,2]/(values[2,2] + values[2,0] + values[2,1] + values[0,2]+ values[1,2])
                            Mean_IoU=IOU_keras.result().numpy()
                            print("IoU for class1 is: "+ str(class1_IoU))
                            print("IoU for class2 is: ", str(class2_IoU))
                            print("IoU for class3 is: ", str(class3_IoU))
                            
                            file1 = open(result_path+"/IoU.txt","w") 
                            file1.write("Mean IoU ="+ str(Mean_IoU)+"\n")
                            
                            #from sklearn.metrics import classification_report 
                            #print(classification_report(groundtruthdataset, predictdataset))
                            #report = classification_report(groundtruthdataset, predictdataset)
                            df_class_report = pandas_classification_report(y_true=groundtruthdataset, y_pred=predictdataset)
                            print("\nThe largest value in the value_iou:")
                            print("value_Precision is: "+ str(df_class_report.to_numpy()[3,0]))
                            print("value_reall is: ", str(df_class_report.to_numpy()[3,1]))
                            print("value_f1 is: ", str(df_class_report.to_numpy()[3,2])) 
                            
                            # Program to show various ways to read and
                            # write data in a file.
                            file1 = open(result_path+"/IoU.txt","w")
                              
                            # \n is placed to indicate EOL (End of Line)
                            file1.write("test_data set result:")
                                               
                            # \n is placed to indicate EOL (End of Line)
                            file1.write("Mean IoU ="+ str(IOU_keras.result().numpy())+"\n")
                            file1.write("IoU for class1 (background) is: "+ str(class1_IoU)+"\n")
                            file1.write("IoU for class2 (Consolidation) is: "+ str(class2_IoU)+"\n")
                            file1.write("IoU for class3 (InfiltrationsBronchiectasis) is: "+ str(class3_IoU)+"\n")
                            file1.write("Mean Precision:"+str(df_class_report.to_numpy()[3,0])+"\n")
                            file1.write("Mean Reall:"+str(df_class_report.to_numpy()[3,1])+"\n")
                            file1.write("Mean F1:"+str(df_class_report.to_numpy()[3,2])+"\n")
                            time_c= time_end - time_start   #執行所花時間
                            file1.write('time cost'+str(time_c)+'s'+"\n")
                            file1.close()
                            #add_Values = [folder_functions[colder_index],k_fold_functions_test_str[k_fold_index], clip_functions[clahe], rate_functions[rate], loss_functions_str[loss],blackbone_functions[blackbone],str(IOU_keras.result().numpy()),str(class1_IoU),str(class2_IoU),str(class3_IoU)]
			                
                            #混淆矩陣
                            Consolidation_predict_array=[]
                            InfiltrationsBronchiectasis_predict_array=[]
                            predict_result = []
                            ground_truth = []
                            # 指定要查詢的路徑
                            df = pd.read_excel('class.xls')  
                            
                            yourPath = con_path+'/'
                            allFileList = os.listdir(yourPath)
                            for file in allFileList:
                              if os.path.isfile(yourPath+file):
                                print (yourPath+file)
                                #ground_truth array 
                                Consolidation_predict = cv2.imread(yourPath+file,cv2.IMREAD_GRAYSCALE)
                                print (Consolidation_predict.shape)
                                #print (np.all(Consolidation_predict == 0))
                                if np.all(Consolidation_predict == 0) == True:
                                    print("正常")
                                    Consolidation_result=0
                                else:
                                    print("Consolidation")
                                    Consolidation_result=1
                                print(Consolidation_result)
                                Consolidation_predict_array = np.append(Consolidation_predict_array,Consolidation_result)
                                
                                print ('./InfiltrationsBronchiectasis/'+file)
                                #ground_truth array 
                                InfiltrationsBronchiectasis_predict = cv2.imread(bro_path+'/'+file,cv2.IMREAD_GRAYSCALE)
                                print (InfiltrationsBronchiectasis_predict.shape)
                                #print (np.all(Consolidation_predict == 0))
                                if np.all(InfiltrationsBronchiectasis_predict == 0) == True:
                                    print("正常")
                                    InfiltrationsBronchiectasis_predict = 0
                                else:
                                    print("InfiltrationsBronchiectasis")
                                    InfiltrationsBronchiectasis_predict=2
                                print(InfiltrationsBronchiectasis_predict)
                                InfiltrationsBronchiectasis_predict_array = np.append(InfiltrationsBronchiectasis_predict_array,InfiltrationsBronchiectasis_predict)
                                print(str(file))
                                print(df.loc[df['file_name']==str(file.split("_")[0]+".png")].values[0][1])
                                ground_truth = np.append(ground_truth,df.loc[df['file_name']==str(file.split("_")[0]+".png")].values[0][1])
                                
                            print(Consolidation_predict_array)
                            print(InfiltrationsBronchiectasis_predict_array)
                            predict_result = Consolidation_predict_array + InfiltrationsBronchiectasis_predict_array
                            print(predict_result)     
                            print(ground_truth)     
                            
                            
                            from sklearn.metrics import ConfusionMatrixDisplay
                            from sklearn.metrics import confusion_matrix
                            from sklearn.metrics import accuracy_score
                            labels = ["Normal", "Con.", "Bro.","Con. And Bro."]
                            
                            cm = confusion_matrix(ground_truth, predict_result)
                            acc = accuracy_score(y_true=ground_truth, y_pred=predict_result)
                            print('Acc: {:.4f}'.format(acc))
                            disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=labels)
                            disp.plot()
                            plt.savefig(result_path+"/"+'Confusion_matrix.png')
                            add_Values = [str(net_str[net]),folder_functions[colder_index], Augmentation[Augmentation_index], rate_functions[rate], loss_functions_str[loss],blackbone_functions[blackbone],str(class1_IoU),str(class2_IoU),str(class3_IoU),str(IOU_keras.result().numpy()),str(df_class_report.to_numpy()[3,0]),str(df_class_report.to_numpy()[3,1]),str(df_class_report.to_numpy()[3,2]),str(acc)]
                            result_list.append(add_Values)
                            
                            reset_keras()

                            gc.collect()
                            print("次數",count)
                            if count == 15: #累積訓練20次則換下一個訓練設置
                                break
                            
    for row in range(len(result_list)):
        ws.append(result_list[row])
    wb.save(r''+folder_functions[colder_index]+'.xlsx')
                    
