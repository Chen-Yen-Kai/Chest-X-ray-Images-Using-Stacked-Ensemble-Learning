# -*- coding: utf-8 -*-
"""
Created on Mon Mar  6 10:33:01 2023

@author: EN308
"""

import pickle
from mlxtend.plotting import plot_decision_regions
from sklearn.metrics import classification_report
import os
#os.environ["CUDA_VISIBLE_DEVICES"]='0' 
os.environ["TF_CPP_MIN_LOG_LEVEL"]= '3'
os.environ["SM_FRAMEWORK"] = "tf.keras"
import Models_psp.models as model_net_psp
from keras_unet_collection import losses as los
#####################################################
from keras_pyramid_pooling_module import PyramidPoolingModule
from tensorflow.keras.models import load_model
import Models.models as model_net
from keras_deeplab_v3_plus_master import model as modelv3
import matplotlib.pyplot as plt
import os
import cv2
from PIL import Image
import numpy as np
import pandas as pd
import time
start = time.time()
from  sklearn.metrics import precision_recall_fscore_support
import Models.resnet50 as resnet50
def pandas_classification_report(y_true, y_pred):
    metrics_summary = precision_recall_fscore_support(
            y_true=y_true, 
            y_pred=y_pred)
    
    avg = list(precision_recall_fscore_support(
            y_true=y_true, 
            y_pred=y_pred,
            average='macro'))

    metrics_sum_index = ['precision', 'recall', 'f1-score', 'support']
    class_report_df = pd.DataFrame(
        list(metrics_summary),
        index=metrics_sum_index)
    
    support = class_report_df.loc['support']
    total = support.sum() 
    avg[-1] = total
    
    class_report_df['avg / total'] = avg

    return class_report_df.T

def set_regularization(model, 
                       kernel_regularizer=None, 
                       bias_regularizer=None, 
                       activity_regularizer=None):
    
    for layer in model.layers:
        
        # set kernel_regularizer
        if kernel_regularizer is not None and hasattr(layer, 'kernel_regularizer'):
            layer.kernel_regularizer = kernel_regularizer

        # set bias_regularizer
        if bias_regularizer is not None and hasattr(layer, 'bias_regularizer'):
            layer.bias_regularizer = bias_regularizer

        # set activity_regularizer
        if activity_regularizer is not None and hasattr(layer, 'activity_regularizer'):
            layer.activity_regularizer = activity_regularizer

# Reset Keras Session
import tensorflow as tf
from tensorflow.compat.v1.keras.backend import clear_session
# def class_differ (argmax):
#     for image_index in range(21):
#         #print("argmax[index].shape")
#         #print(argmax[image_index].shape)
#         prediction_con=np.zeros((512,512,1))
#         prediction_bro=np.zeros((512,512,1))
#         prediction_index_x=np.where(argmax[image_index]==1)[0]
#         #print("prediction_index_x.shape")
#         #print(prediction_index_x.shape)
#         prediction_index_y=np.where(argmax[image_index]==1)[1]
#         #print("prediction_index_y.shape")
#         #print(prediction_index_y.shape)
        
#         for index in range(len(prediction_index_x)):
#             prediction_con[prediction_index_x[index],prediction_index_y[index]]=[255]
#         prediction_index_x=np.where(argmax[image_index]==2)[0]
#         prediction_index_y=np.where(argmax[image_index]==2)[1]
#         for index in range(len(prediction_index_x)):
#             prediction_bro[prediction_index_x[index],prediction_index_y[index]]=[255]
#         #kernel=np.ones((25,25),np.uint8)

#         #prediction_con=cv2.morphologyEx(prediction_con, cv2.MORPH_OPEN, kernel)
#         #prediction_bro=cv2.morphologyEx(prediction_bro, cv2.MORPH_OPEN, kernel)
#         #cv2.imwrite(str(image_index)+"prediction_bro_.png", prediction_bro)
#         #cv2.imwrite(str(image_index)+"prediction_con_.png", prediction_con)
#         image_mask=np.zeros((512,512))
#         #imdex_y=0
#         for y in range(512):
#             #imdex_x=0
#             for x in range(512):
#                 #print (file.split("_"))
#                 #if ((y>=int(file.split("_")[1]))and (y<int(file.split("_")[2]))) and ((x>=int(file.split("_")[3])) and (x<int(file.split("_")[4]))) :
#                    # print(imdex_y,imdex_x)
#                 if prediction_bro[y][x]>0:
#                     image_mask[y][x]=2
#                     #print("prediction_con")
#                 if prediction_con[y][x]>0:
#                     #if prediction_con[imdex_y][imdex_x]>0:
#                     #print("prediction_con")
#                     image_mask[y][x]=1
#                 #imdex_x=imdex_x+1
#             #imdex_y=imdex_y+1  
#         argmax[image_index]=image_mask
#     #cv2.imwrite(predict_path+"/"+file.split("_")[0]+'.png',image_mask)
#     return argmax

def reset_keras():
    sess =  tf.compat.v1.Session()
    clear_session()
    sess.close()
    sess =  tf.compat.v1.Session()

    try:
        pass # this is from global space - change this as you need
    except:
        pass

    #print(gc.collect()) # if it's done something you should see a number being outputted

    # use the same config as you used to create the session
    config = tf.compat.v1.ConfigProto()
    config.gpu_options.per_process_gpu_memory_fraction = 1
    config.gpu_options.visible_device_list = "0"
    tf.compat.v1.keras.backend.set_session(tf.compat.v1.Session(config=config))

def getTestDate (path):
    yourPath = path+"/"
    test_image = []
    allFileList = os.listdir(yourPath)
    for file in allFileList:
      if os.path.isdir(os.path.join(yourPath,file)):
        print("I'm a directory: " + file)
      elif os.path.isfile(yourPath+file):
        #print(file)
        #讀取測試影像
        fimage = cv2.imread(yourPath+file)
        
        image = Image.fromarray(fimage)
        test_image.append(np.array(image))
        # preprocess input
    test_image = np.array(test_image)
    return test_image
def getTestDate_label (path):
    yourPath = path+"/"
    test_image = []
    allFileList = os.listdir(yourPath)
    for file in allFileList:
      if os.path.isdir(os.path.join(yourPath,file)):
        print("I'm a directory: " + file)
      elif os.path.isfile(yourPath+file):
        #讀取測試影像
        fimage = cv2.imread(yourPath+file,0)
        
        image = Image.fromarray(fimage)
        test_image.append(np.array(image))
        # preprocess input
    test_image = np.array(test_image)
    return test_image

def hybrid_loss(y_true, y_pred):

    loss_focal = los.focal_tversky(y_true, y_pred, alpha=0.5, gamma=4/3)
    loss_iou = los.iou_seg(y_true, y_pred)
    
    # (x) 
    #loss_ssim = los.ms_ssim(y_true, y_pred, max_val=1.0, filter_size=4)
    
    return loss_focal + loss_iou #+ loss_ssim

#Set compile=False as we are not loading it for training, only for prediction.

test_path_image = "C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normalLung_segmentation/test_Noneresult_image"

#model1 = load_model("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normal/model/model_U_Net_yun_normal_None_0.0001_total_loss_resnet50.hdf5", compile=False )
val_img1=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normal/val/image_CLAHE")
train_img1=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normal/train/image_CLAHE")
test_imag1=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normal/test_CLAHEresult_image")
print("train_img1_shape:", train_img1.shape)
print("test_imag1_shape:", test_imag1.shape)

clip_functions=["CLAHE"]
blackbone_functions=['densenet121']
net_str=['att_unet']
loss_functions_str =['focal_tversky']
rate_functions =[1e-04] 
folder_functions=["yun_normal"]
for clahe in range(len(clip_functions)):
    for net in range(len(net_str)):  
                print(net_str[net])
                # 载入模型 U-Net
                for blackbone in range(len(blackbone_functions)):
                    print(blackbone_functions[blackbone])
                    for loss in range(len(loss_functions_str)):
                        print(loss_functions_str[loss])
                        for rate in range(len(rate_functions)):
                            print(rate_functions[rate])
                            Mean_IoU=0
                            #建立plot放置資料夾
                            #if blackbone_functions[blackbone] == "none" :
                            #    model = model_net.Unet(use_backbone=False, input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                            #net_str='U_Net','att_unet',"unetplusplus","att_unetplusplus"
                            if blackbone_functions[blackbone] == "none" :
                                pass
                            else:
                                model = model_net.Unet(use_backbone=True,backbone_name=blackbone_functions[blackbone], input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                            
                            model_input = model.input
                            
                            model_output = model.get_layer('final_conv').output
                            from tensorflow.keras.layers import Dropout
                            from tensorflow.keras.models import Model
                            
                            #add dropoutactivations
                            model_output = Dropout(0.3)(model_output)
                            #add activation
                            output =tf.keras.activations.softmax(model_output)
                            model =Model(model.input, output)
                            from tensorflow.keras import regularizers
                            set_regularization(model, kernel_regularizer=regularizers.l2(0.0001),bias_regularizer=regularizers.l2(0.0001))
                            
                            model.summary()
                            
                            #model = sm.utils.set_regularization(model, kernel_regularizer=regularizers.l2(0.001))
        
                            filepath="model_"+str(net_str[net])+"_"+folder_functions[0]+"_"+str(clip_functions[clahe])+"_"+str(rate_functions[rate])+"_"+loss_functions_str[loss]+"_"+blackbone_functions[blackbone]+".hdf5"
                            print(filepath)
     
                            
                            model.load_weights("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/"+folder_functions[0]+'/model/'+filepath)
                            model1=model

#model2 = load_model('C:/Users/EN308/Desktop/Abdiel\2class_MODEL\yun_normal\model\model_att_unet_yun_normal_CLAHE_0.0001_total_loss_resnet50.hdf5', compile=False)
val_img2=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normalLung_segmentation/val/image")
train_img2=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normalLung_segmentation/train/image")
test_imag2=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normalLung_segmentation/test_Noneresult_image")

clip_functions=["None"]
blackbone_functions=['densenet121']
net_str=['unetplusplus']
loss_functions_str =['total_loss']
rate_functions =[1e-04] 
folder_functions=["yun_normalLung_segmentation"]
for clahe in range(len(clip_functions)):
    for net in range(len(net_str)):  
                print(net_str[net])
                # 载入模型 U-Net
                for blackbone in range(len(blackbone_functions)):
                    print(blackbone_functions[blackbone])
                    for loss in range(len(loss_functions_str)):
                        print(loss_functions_str[loss])
                        for rate in range(len(rate_functions)):
                            print(rate_functions[rate])
                            Mean_IoU=0
                            #建立plot放置資料夾
                            #if blackbone_functions[blackbone] == "none" :
                            #    model = model_net.Unet(use_backbone=False, input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                            #net_str='U_Net','att_unet',"unetplusplus","att_unetplusplus"
                            if blackbone_functions[blackbone] == "none" :
                                pass
                            else:
                                model = model_net.Xnet(use_backbone=True,backbone_name=blackbone_functions[blackbone], input_shape=(512,512,3),attention=False,classes=3,activation="softmax")
                            
                            model_input = model.input
                            
                            model_output = model.get_layer('final_conv').output
                            from tensorflow.keras.layers import Dropout
                            from tensorflow.keras.models import Model
                            
                            #add dropoutactivations
                            model_output = Dropout(0.3)(model_output)
                            #add activation
                            output =tf.keras.activations.softmax(model_output)
                            model =Model(model.input, output)
                            from tensorflow.keras import regularizers
                            #set_regularization(model, kernel_regularizer=regularizers.l2(0.0001),bias_regularizer=regularizers.l2(0.0001))
                            model.summary()
                            #model = sm.utils.set_regularization(model, kernel_regularizer=regularizers.l2(0.001))
        
                            filepath="model_"+str(net_str[net])+"_"+folder_functions[0]+"_"+str(clip_functions[clahe])+"_"+str(rate_functions[rate])+"_"+loss_functions_str[loss]+"_"+blackbone_functions[blackbone]+".hdf5"
                            print(filepath)
     
                            
                            model.load_weights("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/"+folder_functions[0]+'/model/'+filepath)
                            model2=model

#model3 = load_model('C:/Users/EN308/Desktop/Abdiel\2class_MODEL\yun_normal\model\model_unetplusplus_yun_normal_None_0.0001_focal_tversky_resnet50.hdf5', compile=False)
val_img3=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_Bone_suppression/val/image")
train_img3=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_Bone_suppression/train/image")
test_imag3=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_Bone_suppression/test_Noneresult_image")

clip_functions=["None"]
blackbone_functions=['resnet50']
net_str=['att_unetplusplus']
loss_functions_str =['focal_tversky']
rate_functions =[1e-04] 
folder_functions=["yun_Bone_suppression"]
for clahe in range(len(clip_functions)):
    for net in range(len(net_str)):  
                print(net_str[net])
                # 载入模型 U-Net
                for blackbone in range(len(blackbone_functions)):
                    print(blackbone_functions[blackbone])
                    for loss in range(len(loss_functions_str)):
                        print(loss_functions_str[loss])
                        for rate in range(len(rate_functions)):
                            print(rate_functions[rate])
                            Mean_IoU=0
                            #建立plot放置資料夾
                            #if blackbone_functions[blackbone] == "none" :
                            #    model = model_net.Unet(use_backbone=False, input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                            #net_str='U_Net','att_unet',"unetplusplus","att_unetplusplus"
                            if blackbone_functions[blackbone] == "none" :
                                pass
                            else:
                                model = model_net.Xnet(use_backbone=True,backbone_name=blackbone_functions[blackbone], input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                            
                            model_input = model.input
                            
                            model_output = model.get_layer('final_conv').output
                            from tensorflow.keras.layers import Dropout
                            from tensorflow.keras.models import Model
                            
                            #add dropoutactivations
                            model_output = Dropout(0.3)(model_output)
                            #add activation
                            output =tf.keras.activations.softmax(model_output)
                            model =Model(model.input, output)
                            from tensorflow.keras import regularizers
                            set_regularization(model, kernel_regularizer=regularizers.l2(0.0001),bias_regularizer=regularizers.l2(0.0001))
                            model.summary()
                            #model = sm.utils.set_regularization(model, kernel_regularizer=regularizers.l2(0.001))
        
                            filepath="model_"+str(net_str[net])+"_"+folder_functions[0]+"_"+str(clip_functions[clahe])+"_"+str(rate_functions[rate])+"_"+loss_functions_str[loss]+"_"+blackbone_functions[blackbone]+".hdf5"
                            print(filepath)
     
                            
                            model.load_weights("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/"+folder_functions[0]+'/model/'+filepath)
                            model3=model

#model4 = load_model('C:/Users/EN308/Desktop/Abdiel\2class_MODEL\yun_Bone_suppression\model\model_att_unetplusplus_yun_Bone_suppression_None_0.0001_focal_tversky_resnet50.hdf5', compile=False)
val_img4=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normal/val/image")
train_img4=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normal/train/image")
test_imag4=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normal/test_Noneresult_image")

clip_functions=["None"]
blackbone_functions=['densenet121']
net_str=['att_unetplusplus_psp']
loss_functions_str =['categorical_crossentropy']
rate_functions =[1e-04] 
folder_functions=["yun_normal"]
for clahe in range(len(clip_functions)):
    for net in range(len(net_str)):  
                print(net_str[net])
                # 载入模型 U-Net
                for blackbone in range(len(blackbone_functions)):
                    print(blackbone_functions[blackbone])
                    for loss in range(len(loss_functions_str)):
                        print(loss_functions_str[loss])
                        for rate in range(len(rate_functions)):
                            print(rate_functions[rate])
                            Mean_IoU=0
                            #建立plot放置資料夾
                            #if blackbone_functions[blackbone] == "none" :
                            #    model = model_net.Unet(use_backbone=False, input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                            #net_str='U_Net','att_unet',"unetplusplus","att_unetplusplus"
                            if blackbone_functions[blackbone] == "none" :
                                pass
                            else:
                                model = model_net_psp.Xnet(use_backbone=True,backbone_name=blackbone_functions[blackbone], input_shape=(512,512,3),attention=True,classes=3,activation="softmax",decoder_block_type = "transpose")
                            
                            model_input = model.input
                            
                            model_output = model.get_layer('final_conv').output
                            from tensorflow.keras.layers import Dropout
                            from tensorflow.keras.models import Model
                            
                            #add dropoutactivations
                            model_output = Dropout(0.3)(model_output)
                            #add activation
                            output =tf.keras.activations.softmax(model_output)
                            model =Model(model.input, output)
                            from tensorflow.keras import regularizers
                            set_regularization(model, kernel_regularizer=regularizers.l2(0.0001),bias_regularizer=regularizers.l2(0.0001))
                            model.summary()
                            #model = sm.utils.set_regularization(model, kernel_regularizer=regularizers.l2(0.001))
        
                            filepath="model_"+str(net_str[net])+"_"+folder_functions[0]+"_"+str(clip_functions[clahe])+"_"+str(rate_functions[rate])+"_"+loss_functions_str[loss]+"_"+blackbone_functions[blackbone]+".hdf5"
                            print(filepath)
     
                            
                            model.load_weights("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/"+folder_functions[0]+'/model/'+filepath)
                            model4=model

val_img5=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normalLung_segmentation/val/image")
train_img5=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normalLung_segmentation/train/image")
test_imag5=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normalLung_segmentation/test_Noneresult_image")

clip_functions=["None"]
blackbone_functions=['resnet50']
net_str=['att_unetplusplus_psp']
loss_functions_str =['total_loss']
rate_functions =[1e-04] 
folder_functions=["yun_normalLung_segmentation"]
for clahe in range(len(clip_functions)):
    for net in range(len(net_str)):  
                print(net_str[net])
                # 载入模型 U-Net
                for blackbone in range(len(blackbone_functions)):
                    print(blackbone_functions[blackbone])
                    for loss in range(len(loss_functions_str)):
                        print(loss_functions_str[loss])
                        for rate in range(len(rate_functions)):
                            print(rate_functions[rate])
                            Mean_IoU=0
                            #建立plot放置資料夾
                            #if blackbone_functions[blackbone] == "none" :
                            #    model = model_net.Unet(use_backbone=False, input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                            #net_str='U_Net','att_unet',"unetplusplus","att_unetplusplus"
                            if blackbone_functions[blackbone] == "none" :
                                pass
                            else:
                                model = model_net_psp.Xnet(use_backbone=True,backbone_name=blackbone_functions[blackbone], input_shape=(512,512,3),attention=True,classes=3,activation="softmax",decoder_block_type = "transpose")
                            
                            model_input = model.input
                            
                            model_output = model.get_layer('final_conv').output
                            from tensorflow.keras.layers import Dropout
                            from tensorflow.keras.models import Model
                            
                            #add dropoutactivations
                            model_output = Dropout(0.3)(model_output)
                            #add activation
                            output =tf.keras.activations.softmax(model_output)
                            model =Model(model.input, output)
                            from tensorflow.keras import regularizers
                            set_regularization(model, kernel_regularizer=regularizers.l2(0.0001),bias_regularizer=regularizers.l2(0.0001))
                            model.summary()
                            #model = sm.utils.set_regularization(model, kernel_regularizer=regularizers.l2(0.001))
        
                            filepath="model_"+str(net_str[net])+"_"+folder_functions[0]+"_"+str(clip_functions[clahe])+"_"+str(rate_functions[rate])+"_"+loss_functions_str[loss]+"_"+blackbone_functions[blackbone]+".hdf5"
                            print(filepath)
     
                            
                            model.load_weights("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/"+folder_functions[0]+'/model/'+filepath)
                            model5=model

val_img6=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_Bone_suppression_Lung_segmentation/val/image")
train_img6=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_Bone_suppression_Lung_segmentation/train/image")
test_imag6=getTestDate("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_Bone_suppression_Lung_segmentation/test_Noneresult_image")

clip_functions=["None"]
blackbone_functions=['densenet121']
net_str=['unetplusplus']
loss_functions_str =['hybrid_loss']
rate_functions =[1e-04] 
folder_functions=["yun_Bone_suppression_Lung_segmentation"]
for clahe in range(len(clip_functions)):
    for net in range(len(net_str)):  
                print(net_str[net])
                # 载入模型 U-Net
                for blackbone in range(len(blackbone_functions)):
                    print(blackbone_functions[blackbone])
                    for loss in range(len(loss_functions_str)):
                        print(loss_functions_str[loss])
                        for rate in range(len(rate_functions)):
                            print(rate_functions[rate])
                            Mean_IoU=0
                            #建立plot放置資料夾
                            #if blackbone_functions[blackbone] == "none" :
                            #    model = model_net.Unet(use_backbone=False, input_shape=(512,512,3),attention=True,classes=3,activation="softmax")
                            #net_str='U_Net','att_unet',"unetplusplus","att_unetplusplus"
                            if blackbone_functions[blackbone] == "none" :
                                pass
                            else:
                                model = model_net.Xnet(use_backbone=True,backbone_name=blackbone_functions[blackbone], input_shape=(512,512,3),attention=False,classes=3,activation="softmax")
                            
                            model_input = model.input
                            
                            model_output = model.get_layer('final_conv').output
                            from tensorflow.keras.layers import Dropout
                            from tensorflow.keras.models import Model
                            
                            #add dropoutactivations
                            model_output = Dropout(0.3)(model_output)
                            #add activation
                            output =tf.keras.activations.softmax(model_output)
                            model =Model(model.input, output)
                            from tensorflow.keras import regularizers
                            #set_regularization(model, kernel_regularizer=regularizers.l2(0.0001),bias_regularizer=regularizers.l2(0.0001))
                            model.summary()
                            #model = sm.utils.set_regularization(model, kernel_regularizer=regularizers.l2(0.001))
        
                            filepath="model_"+str(net_str[net])+"_"+folder_functions[0]+"_"+str(clip_functions[clahe])+"_"+str(rate_functions[rate])+"_"+loss_functions_str[loss]+"_"+blackbone_functions[blackbone]+".hdf5"
                            print(filepath)
     
                            
                            model.load_weights("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/"+folder_functions[0]+'/model/'+filepath)
                            model6=model

label_train=getTestDate_label("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normal/train/label") #引入訓練集
label_val=getTestDate_label("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normal/val/label")
label=getTestDate_label("C:/Users/EN308/Desktop/Abdiel/2class_MODEL/yun_normal/test_label")
print("train_label_shape:", label_train.shape)
print("test_label_shape:", label.shape)

#Weighted average ensemble
models = [model1, model2, model3, model4, model5]
preds = []

pred1_add = []
pred2_add = []
pred3_add = []
pred4_add = []
pred5_add = []
pred6_add = []
import openpyxl
# 儲存 Excel 活頁簿至檔案
# 建立 Excel 活頁簿
xlsx_file='lr_ALL.xlsx'
workbook = openpyxl.Workbook()
#workbook.save(xlsx_file)   #儲存
workbook.save(xlsx_file)
result_list = [["Mean IoU","Mean precision","Mean recall","Mean f1-score","Acc","Kernel Size"]]
wb=openpyxl.load_workbook(r''+"lr_ALL"+'.xlsx')
# 新建一個sheet的物件
ws = wb.create_sheet(title='Sheet1',index=0)
#獲取行數和列數：
rows = ws.max_row   #獲取行數
cols = ws.max_column    #獲取列數


#測試影像
for index in range(21):   
    test_imag1_index=test_imag1[index:index+1,:,:,:]
    test_imag2_index=test_imag2[index:index+1,:,:,:]
    test_imag3_index=test_imag3[index:index+1,:,:,:]
    test_imag4_index=test_imag4[index:index+1,:,:,:]
    test_imag5_index=test_imag5[index:index+1,:,:,:]
    test_imag6_index=test_imag6[index:index+1,:,:,:]
    test_imag1_index = np.array(test_imag1_index)
    test_imag2_index = np.array(test_imag2_index)
    test_imag3_index = np.array(test_imag3_index)
    test_imag4_index = np.array(test_imag4_index)
    test_imag5_index = np.array(test_imag5_index)
    test_imag6_index = np.array(test_imag6_index)
    
    pred1 = model1.predict(test_imag1_index)
    reset_keras()
    pred2 = model2.predict(test_imag2_index)
    reset_keras()
    pred3 = model3.predict(test_imag3_index)
    reset_keras()
    pred4 = model4.predict(test_imag4_index)
    reset_keras()
    pred5 = model5.predict(test_imag5_index)
    reset_keras()
    pred6 = model6.predict(test_imag6_index)
    reset_keras()
    pred1_add.append(np.array(pred1[0,:,:,:]))
    pred2_add.append(np.array(pred2[0,:,:,:]))
    pred3_add.append(np.array(pred3[0,:,:,:]))
    pred4_add.append(np.array(pred4[0,:,:,:]))
    pred5_add.append(np.array(pred5[0,:,:,:]))
    pred6_add.append(np.array(pred6[0,:,:,:]))

#先分類後回歸
pred1_add = np.argmax(np.array(pred1_add), axis=3)[:,:,:]
pred2_add = np.argmax(np.array(pred2_add), axis=3)[:,:,:]
pred3_add = np.argmax(np.array(pred3_add), axis=3)[:,:,:]
pred4_add = np.argmax(np.array(pred4_add), axis=3)[:,:,:]
pred5_add = np.argmax(np.array(pred5_add), axis=3)[:,:,:]
pred6_add = np.argmax(np.array(pred6_add), axis=3)[:,:,:]
pred1_add = pred1_add.reshape((pred1_add.shape[0], pred1_add.shape[1], pred1_add.shape[2], 1))
pred2_add = pred2_add.reshape((pred2_add.shape[0], pred2_add.shape[1], pred2_add.shape[2], 1))
pred3_add = pred3_add.reshape((pred3_add.shape[0], pred3_add.shape[1], pred3_add.shape[2], 1))
pred4_add = pred4_add.reshape((pred4_add.shape[0], pred4_add.shape[1], pred4_add.shape[2], 1))
pred5_add = pred5_add.reshape((pred5_add.shape[0], pred5_add.shape[1], pred5_add.shape[2], 1))
pred6_add = pred6_add.reshape((pred6_add.shape[0], pred6_add.shape[1], pred6_add.shape[2], 1))
X_test = np.concatenate([pred1_add, pred2_add, pred3_add, pred4_add, pred6_add], axis=3).astype(float)



result = np.zeros((21,512,512))
print(result.shape)

for i in range(21):
    print(i)
    for j in range(512):
        for k in range(512):
            if pred1_add[i,j,k]==pred2_add[i,j,k]==pred3_add[i,j,k]==pred4_add[i,j,k]==pred6_add[i,j,k]:  #AND
                result[i,j,k] = pred1_add[i,j,k]  #AND
            
            #or
            # num_list=[]
            # num_list = X_test[i,j,k,:]
            # num_list = num_list.tolist()
            # a = num_list.count(0)
            # b = num_list.count(1)
            # c = num_list.count(2)
            # result[i,j,k] = np.argmax([a,b,c])
    

ensemble_prediction=np.array(result)
#ensemble_prediction=ensemble_prediction.reshape((21,512,512))


#SIZE=[0, 5, 10, 15, 20, 25]
SIZE=[0]
for SIZE_i in SIZE:
    result_path = "./emable_result/result_path"+str(SIZE_i)
    plot_path =  result_path+"/plot_path"
    predict_path = result_path+"/predict"
    con_path = result_path+"/con_path"
    bro_path = result_path+"/bro_path"
    
    yourPath = test_path_image+"/"
    allFileList = os.listdir(yourPath)
    result_index = 0
    for file in allFileList:
      if os.path.isdir(os.path.join(yourPath,file)):
        print("I'm a directory: " + file)
      elif os.path.isfile(yourPath+file):
        print(file)
        prediction_array=np.zeros((512,512,3))
        prediction_con=np.zeros((512,512,1))
        prediction_bro=np.zeros((512,512,1))
        prediction_index_x=np.where(ensemble_prediction[result_index]==1)[0]
        prediction_index_y=np.where(ensemble_prediction[result_index]==1)[1]
        for index in range(len(prediction_index_x)):
            prediction_array[prediction_index_x[index],prediction_index_y[index]]=[0,0,255]
            prediction_con[prediction_index_x[index],prediction_index_y[index]]=[255]
        #cv2.imwrite(con_path+"/"+file[result_index], prediction_con)
       
        prediction_index_x=np.where(ensemble_prediction[result_index]==2)[0]
        prediction_index_y=np.where(ensemble_prediction[result_index]==2)[1]
        for index in range(len(prediction_index_x)):
            prediction_array[prediction_index_x[index],prediction_index_y[index]]=[0,255,255]
            prediction_bro[prediction_index_x[index],prediction_index_y[index]]=[255]
        #cv2.imwrite(bro_path+"/"+file[result_index], prediction_bro)
        
        #開運算
        kernel=np.ones((SIZE_i,SIZE_i),np.uint8)
        prediction_con=cv2.morphologyEx(prediction_con, cv2.MORPH_OPEN, kernel)
        prediction_bro=cv2.morphologyEx(prediction_bro, cv2.MORPH_OPEN, kernel)
               
        
        cv2.imwrite(bro_path+"/"+file, prediction_bro)
        cv2.imwrite(con_path+"/"+file, prediction_con)
        prediction_con = cv2.resize(prediction_con, (int(file.split("_")[5]),int(file.split("_")[5])), interpolation=cv2.INTER_AREA)
        prediction_bro = cv2.resize(prediction_bro, (int(file.split("_")[5]),int(file.split("_")[5])), interpolation=cv2.INTER_AREA)
        #cv2.imwrite(file[result_index].split("_")[0]+'prediction_con.png', prediction_con)
        #cv2.imwrite(file[result_index].split("_")[0]+'prediction_bro.png', prediction_bro)
   
        #解答繪製
        groundtruth_image = cv2.imread("./groundtruth/"+file.split("_")[0]+".png")
        groundtruth_array=np.zeros((groundtruth_image.shape[0],groundtruth_image.shape[1],3))
        groundtruth_index_x=np.where(groundtruth_image==1)[0]
        groundtruth_index_y=np.where(groundtruth_image==1)[1]
        for index in range(len(groundtruth_index_x)):
            groundtruth_array[groundtruth_index_x[index],groundtruth_index_y[index]]=[0,0,255]
        groundtruth_index_x=np.where(groundtruth_image==2)[0]
        groundtruth_index_y=np.where(groundtruth_image==2)[1]
        for index in range(len(groundtruth_index_x)):
            groundtruth_array[groundtruth_index_x[index],groundtruth_index_y[index]]=[0,255,255]
        cv2.imwrite(result_path+"/"+file.split("_")[0]+'groundtruth.png', groundtruth_array)
        #回到原始圖像
        #回到原始圖像
        #回到原始圖像
        or_size_image = cv2.imread("./image_orginal_marked/"+file.split("_")[0]+".jpg")
        prediction_bro=np.where(prediction_bro<1,0,2)
        prediction_con=np.where(prediction_con<1,0,1)
        allsize=prediction_con+prediction_bro
        allsize[allsize==3]=1
       
        #allsize=np.pad(allsize, ((int(file.split("_")[3]), int(file.split("_")[1])),(or_size_image.shape[0]-int(file.split("_")[2]), or_size_image.shape[1]-int(file.split("_")[4]))), 'constant', constant_values=((0, 0)))
        allsize=np.pad(allsize, ((int(file.split("_")[1]),or_size_image.shape[0]-int(file.split("_")[2])),(int(file.split("_")[3]),or_size_image.shape[1]-int(file.split("_")[4]))), 'constant', constant_values=(0, 0))
       
        allsize_rgb_array=np.zeros((groundtruth_image.shape[0],groundtruth_image.shape[1],3))
        allsize_rgb=np.zeros((allsize_rgb_array.shape[0],allsize_rgb_array.shape[1],3))
        allsize_rgb_x=np.where(allsize==1)[0]
        allsize_rgb_y=np.where(allsize==1)[1]
        for index in range(len(allsize_rgb_x)):
            allsize_rgb_array[allsize_rgb_x[index],allsize_rgb_y[index]]=[0,0,255]
        allsize_rgb_x=np.where(allsize==2)[0]
        allsize_rgb_y=np.where(allsize==2)[1]
        for index in range(len(allsize_rgb_x)):
            allsize_rgb_array[allsize_rgb_x[index],allsize_rgb_y[index]]=[0,255,255]

        cv2.imwrite(predict_path+"/"+file.split("_")[0]+'.png', allsize)
        cv2.imwrite(result_path+"/"+file.split("_")[0]+'image_mask_rgb.png', allsize_rgb_array)      
        prediction = cv2.imread(result_path+"/"+file.split("_")[0]+'image_mask_rgb.png')
        result = cv2.addWeighted(prediction, 0.3, or_size_image, 1, 0, or_size_image)
        cv2.imwrite(result_path+"/"+file.split("_")[0]+'.png', result)
        #下一張
        result_index+=1
        
        #圖表顯示
        fig = plt.figure(figsize=(80,40))
        image_or=cv2.imread("./image_orginal/"+file.split("_")[0]+'.png')
        ax = fig.add_subplot(1, 4, 1)
        image_or_plot=cv2.cvtColor(image_or,cv2.COLOR_BGR2RGB)
        imgplot = plt.imshow(image_or_plot)
        ax.set_title('Original image',fontsize=40)
        plt.axis('off')
        
        ax = fig.add_subplot(1, 4, 2)
        Ground_truth = cv2.imread(result_path+"/"+file.split("_")[0]+'groundtruth.png')
        Ground_truth_plot=cv2.cvtColor(Ground_truth,cv2.COLOR_BGR2RGB)
        imgplot = plt.imshow(Ground_truth_plot)
        ax.set_title('Ground truth',fontsize=40)
        plt.axis('off')
        
        ax = fig.add_subplot(1, 4, 3)
        prediction = cv2.imread(result_path+"/"+file.split("_")[0]+'image_mask_rgb.png')
        prediction_plot=cv2.cvtColor(prediction,cv2.COLOR_BGR2RGB)
        imgplot = plt.imshow(prediction_plot)
        ax.set_title('Prediction mask',fontsize=40)
        plt.axis('off')
        
        ax = fig.add_subplot(1, 4, 4)
        overlap_reault = cv2.imread(result_path+"/"+file.split("_")[0]+'.png')
        overlap_reault_plot=cv2.cvtColor(overlap_reault,cv2.COLOR_BGR2RGB)
        imgplot = plt.imshow(overlap_reault_plot)
        #mark_reault_plot=cv2.cvtColor(mark,cv2.COLOR_BGR2RGB)
        #imgplot = plt.imshow(mark_reault_plot)
        ax.set_title('Overlapping results',fontsize=40)
        plt.axis('off')
            
        plt.savefig(plot_path+"/"+file[:-4]+'_plot.png')
        plt.rcParams.update({'figure.max_open_warning': 0})
        plt.close(fig)
       
    predictdirectory=predict_path+"/"
    predictdataset=np.array([])
    groundtruthdataset=np.array([])
    masks = os.listdir(predictdirectory)
    for i, image_name in enumerate(masks):
        if (image_name.split('.')[1] == 'png'):
            predict_image = cv2.imread(predictdirectory+image_name,0)
            groundtruth_image = cv2.imread("./groundtruth/"+image_name,0)
            print("./groundtruth/"+image_name)
            print(predictdirectory+image_name)
            print(predict_image.shape)
            print(groundtruth_image.shape) 
            height,width=predict_image.shape
            predict_image=np.reshape(predict_image, (height*width))
            groundtruth_image=np.reshape(groundtruth_image, (height*width))
            predictdataset = np.concatenate((predictdataset, predict_image), axis = 0)
            groundtruthdataset = np.concatenate((groundtruthdataset, groundtruth_image), axis = 0)
   
    print(predictdataset.shape)
    print(groundtruthdataset.shape) 
   
       
    from keras.metrics import MeanIoU
    #
    n_classes = 3
    IOU_keras = MeanIoU(num_classes=n_classes)  
    #predictdataset = np.array(predictdataset)
    #groundtruthdataset = np.array(groundtruthdataset)
    print("predictdataset.shape")
    print(predictdataset)
    print(groundtruthdataset)
    IOU_keras.update_state(groundtruthdataset, predictdataset)
    print("Mean IoU =", IOU_keras.result().numpy())
   
    print(classification_report(groundtruthdataset, predictdataset))
    report = classification_report(groundtruthdataset, predictdataset)
    df_class_report = pandas_classification_report(y_true=groundtruthdataset, y_pred=predictdataset)
    print("\nThe largest value in the value_iou:")
    print("value_Precision is: "+ str(df_class_report.to_numpy()[3,0]))
    print("value_reall is: ", str(df_class_report.to_numpy()[3,1]))
    print("value_f1 is: ", str(df_class_report.to_numpy()[3,2])) 
   
    #To calculate I0U for each class...
    values = np.array(IOU_keras.get_weights()).reshape(n_classes, n_classes)
    print(values)
    class1_IoU = values[0,0]/(values[0,0] + values[0,1] + values[0,2] + values[1,0]+ values[2,0])
    class2_IoU = values[1,1]/(values[1,1] + values[1,0] + values[1,2] + values[0,1]+ values[2,1])
    class3_IoU = values[2,2]/(values[2,2] + values[2,0] + values[2,1] + values[0,2]+ values[1,2])
    Mean_IoU=IOU_keras.result().numpy()
    print("IoU for class1 is: "+ str(class1_IoU))
    print("IoU for class2 is: ", str(class2_IoU))
    print("IoU for class3 is: ", str(class3_IoU))
   
   
   
    #混淆矩陣
    print("混淆矩陣")
    Consolidation_predict_array=[]
    InfiltrationsBronchiectasis_predict_array=[]
    predict_result = []
    ground_truth = []
    # 指定要查詢的路徑
    df = pd.read_excel('C:/Users/EN308/Desktop/Abdiel/2class_MODEL/ensemble/class.xls')  
   
    #print(df.at[1,'class'])
   
    yourPath = con_path+'/'
    allFileList = os.listdir(yourPath)
    for file_name in allFileList:
      if os.path.isfile(yourPath+file_name):
        print (yourPath+file_name)
        #ground_truth array 
        Consolidation_predict = cv2.imread(yourPath+file_name,cv2.IMREAD_GRAYSCALE)
        print (Consolidation_predict.shape)
        #print (np.all(Consolidation_predict == 0))
        if np.all(Consolidation_predict == 0) == True:
            print("正常")
            Consolidation_result=0
        else:
            print("Consolidation")
            Consolidation_result=1
        print(Consolidation_result)
        Consolidation_predict_array = np.append(Consolidation_predict_array,Consolidation_result)
       
        print ('./InfiltrationsBronchiectasis/'+file_name)
        #ground_truth array 
        InfiltrationsBronchiectasis_predict = cv2.imread(bro_path+'/'+file_name,cv2.IMREAD_GRAYSCALE)
        print (InfiltrationsBronchiectasis_predict.shape)
        #print (np.all(Consolidation_predict == 0))
        if np.all(InfiltrationsBronchiectasis_predict == 0) == True:
            print("正常")
            InfiltrationsBronchiectasis_predict = 0
        else:
            print("InfiltrationsBronchiectasis")
            InfiltrationsBronchiectasis_predict=2
        print(InfiltrationsBronchiectasis_predict)
        InfiltrationsBronchiectasis_predict_array = np.append(InfiltrationsBronchiectasis_predict_array,InfiltrationsBronchiectasis_predict)
        print(str(file_name))
        print(df.loc[df['file_name']==str(file_name.split("_")[0]+".png")].values[0][1])
        ground_truth = np.append(ground_truth, df.loc[df['file_name']==str(file_name.split("_")[0]+".png")].values[0][1])
       
    print(Consolidation_predict_array)
    print(InfiltrationsBronchiectasis_predict_array)
    predict_result = Consolidation_predict_array + InfiltrationsBronchiectasis_predict_array
    print(predict_result)     
    print(ground_truth)     
   
   
    from sklearn.metrics import ConfusionMatrixDisplay
    from sklearn.metrics import confusion_matrix
    from sklearn.metrics import accuracy_score
    labels = ["Normal", "Con.", "Bro.","Con. And Bro."]
   
    cm = confusion_matrix(ground_truth, predict_result)
    acc = accuracy_score(y_true=ground_truth, y_pred=predict_result)
    print('Acc: {:.4f}'.format(acc))
    disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=labels)
    disp.plot()
    disp.ax_.set(xlabel='Predicted labels', ylabel='True labels')
    from matplotlib import pyplot as plt
    plt.savefig(result_path+"/"+str(SIZE_i)+'_'+'Confusion_matrix.png', dpi=300, bbox_inches='tight', pad_inches=0.0)
   
    add_Values = [str(IOU_keras.result().numpy()),str(df_class_report.to_numpy()[3,0]),str(df_class_report.to_numpy()[3,1]),str(df_class_report.to_numpy()[3,2]),str(acc),str(SIZE_i)]
    #add_Values = [str(IOU_keras.result().numpy()),str(df_class_report.to_numpy()[3,0]),str(df_class_report.to_numpy()[3,1]),str(df_class_report.to_numpy()[3,2]),str(SIZE_i)]
    result_list.append(add_Values)

for row in range(len(result_list)):
    ws.append(result_list[row])
wb.save(r''+"lr_ALL"+'.xlsx')


end = time.time()
print("執行時間：%f 秒" % (end - start))